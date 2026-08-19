import pandas as pd
from config import CONFIG
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =========================================================
# FINAL STRUCTURE (ONLY ONE IR COLUMN)
# =========================================================
DETAIL_COLUMNS = [
    "Date",
    "Module",
    "Point",
    "Voltage",
    "IR(mΩ)",
    "Status"
]

SUMMARY_COLUMNS = [
    "Date",
    "Time",
    "Shift",
    "Operator",
    "Module",
    "Result",
    "Issue"
]

# =========================================================
# SAVE FUNCTION
# =========================================================
def save(summary, details):

    file = CONFIG["excel_file"]

    # ================= SUMMARY =================
    try:
        summary_df = pd.read_excel(file, sheet_name="Summary")
    except:
        summary_df = pd.DataFrame(columns=SUMMARY_COLUMNS)

    summary_df = pd.concat(
        [summary_df, pd.DataFrame([summary])],
        ignore_index=True
    )

    # ================= DETAIL =================
    try:
        detail_df = pd.read_excel(file, sheet_name="Detail")
    except:
        detail_df = pd.DataFrame(columns=DETAIL_COLUMNS)

    if details:

        today = datetime.now().strftime("%d-%m-%Y")

        rows = []

        for row in details:
            rows.append({
                "Date": today,
                "Module": row.get("Module", ""),
                "Point": row.get("Point", ""),
                "Voltage": row.get("Voltage", ""),
                "IR(mΩ)": row.get("IR", ""),   # ONLY IR (NO IR(mΩ))
                "Status": row.get("Status", "")
            })

        new_rows = pd.DataFrame(rows)

        # FORCE CLEAN STRUCTURE (REMOVE OLD IR(mΩ))
        detail_df = detail_df.reindex(columns=DETAIL_COLUMNS, fill_value="")

        detail_df = pd.concat(
            [detail_df, new_rows],
            ignore_index=True
        )

        # blank row after every 15 modules
        module_count = len(detail_df[detail_df["Point"] == "Module"])

        if module_count > 0 and module_count % 15 == 0:
            blank = {col: "" for col in DETAIL_COLUMNS}
            detail_df = pd.concat([detail_df, pd.DataFrame([blank])], ignore_index=True)

    # FINAL COLUMN LOCK
    detail_df = detail_df[DETAIL_COLUMNS]

    # ================= WRITE =================
    with pd.ExcelWriter(file, engine="openpyxl", mode="w") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        detail_df.to_excel(writer, sheet_name="Detail", index=False)

    apply_formatting(file)
    apply_borders(file)
    create_report(file)

# =========================================================
# PASS / FAIL COLORING
# =========================================================
def apply_formatting(file):

    wb = load_workbook(file)
    ws = wb["Detail"]

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    headers = [c.value for c in ws[1]]
    status_col = headers.index("Status") + 1

    for r in range(2, ws.max_row + 1):

        value = ws.cell(r, status_col).value

        if value == "PASS":
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = green

        elif value == "FAIL":
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = red

    wb.save(file)

# =========================================================
# BORDERS + FORMAT + AUTO WIDTH
# =========================================================
def apply_borders(file):

    wb = load_workbook(file)
    ws = wb["Detail"]

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [cell.value for cell in ws[1]]

    module_col = headers.index("Module") + 1

    for r in range(2, ws.max_row + 1):

        if ws.cell(r, module_col).value not in ("", None):

            for c in range(1, len(headers) + 1):
                cell = ws.cell(r, c)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # HEADER STYLE
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # AUTO WIDTH
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_len + 3

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(file)

# =========================================================
# REPORT
# =========================================================
def create_report(file):

    wb = load_workbook(file)

    if "Report" in wb.sheetnames:
        del wb["Report"]

    ws = wb.create_sheet("Report")

    df = pd.read_excel(file, sheet_name="Detail")

    if df.empty:
        wb.save(file)
        return

    total = len(df[df["Point"] == "Module"])
    passed = len(df[(df["Point"] == "Module") & (df["Status"] == "PASS")])
    failed = len(df[(df["Point"] == "Module") & (df["Status"] == "FAIL")])

    yield_percent = round((passed / total) * 100, 2) if total else 0

    ws["A1"] = "DAILY PRODUCTION REPORT"

    ws["A3"] = "Total Modules"
    ws["B3"] = total

    ws["A4"] = "PASS"
    ws["B4"] = passed

    ws["A5"] = "FAIL"
    ws["B5"] = failed

    ws["A7"] = "Yield (%)"
    ws["B7"] = yield_percent

    wb.save(file)